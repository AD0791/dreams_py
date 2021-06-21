from agyw import AgywPrev, AgywPrevCommune
from re import sub
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.borders import Border, Side

# constante openpyxl design reusable elements
my_blue = Color(rgb="0000FF")
blue_fill = PatternFill(patternType="solid",fgColor=my_blue)
my_red = Color(rgb="FF0000")
red_fill = PatternFill(patternType="solid",fgColor=my_red)
bold_12 = Font(size=12,bold=True)
text_datim_title = Alignment(horizontal="center",
                            vertical="bottom",
                            wrapText=True)
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Name Handler
def name_handler(s):
    """This function will Replace all runs of whitespace with a single dash"""
    s = sub(r"[^\w\s]", '', s)
    s = sub(r"\s+", '-', s)
    return s


# init workbook
def init_wb():
    """This function will be responsible for initiate the workbook"""
    wb = Workbook()
    return wb

# init the datim
def init_datim():
    """This function will be responsible for initiate the datim instance"""
    datim = AgywPrev()
    return datim

# Active the workbook 
def active_worksheet():
    """This function will be responsible """
    wb = init_wb
    datim = init_datim
    ws = wb.active
    ws.title = datim.who_am_i
    return ws

# Active 



if __name__ == "__main__":
    init_wb()
    init_datim()
    active_worksheet()